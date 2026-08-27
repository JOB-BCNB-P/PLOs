from openpyxl import load_workbook
p='/home/ubuntu/upload/Curriculummapping-2565วพบกรุงเทพพฤษภาคม65.xlsx'
ws=load_workbook(p,read_only=True,data_only=True).active
for i,row in enumerate(ws.iter_rows(min_row=1,max_row=15,values_only=True), start=1):
    values=[str(v).strip() for v in row if v is not None and str(v).strip()]
    print(i, values[:12])
