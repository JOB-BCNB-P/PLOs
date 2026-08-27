from pathlib import Path
from zipfile import ZipFile
import json
import re
import xml.etree.ElementTree as ET
from openpyxl import load_workbook

UPLOAD = Path('/home/ubuntu/upload')
OUT = Path('/home/ubuntu/plo-assessment-system/docs/source_analysis')
OUT.mkdir(parents=True, exist_ok=True)

files = sorted(UPLOAD.iterdir())
manifest = []
for path in files:
    if not path.is_file():
        continue
    item = {'filename': path.name, 'size_bytes': path.stat().st_size}
    if path.suffix.lower() in {'.xlsx', '.xlsm'}:
        wb = load_workbook(path, read_only=True, data_only=True)
        sheets = []
        for ws in wb.worksheets:
            rows = []
            for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 20), values_only=True):
                rows.append([str(v).strip() if v is not None else '' for v in row[:30]])
            sheets.append({'title': ws.title, 'max_row': ws.max_row, 'max_column': ws.max_column, 'sample_rows': rows})
        item['type'] = 'xlsx'
        item['sheets'] = sheets
    elif path.suffix.lower() == '.docx':
        with ZipFile(path) as zf:
            xml = zf.read('word/document.xml')
        root = ET.fromstring(xml)
        ns = {'w': 'http://schemas.openxmlformats.org/wordprocessingml/2006/main'}
        paragraphs = []
        for p in root.findall('.//w:p', ns):
            text = ''.join(t.text or '' for t in p.findall('.//w:t', ns)).strip()
            if text:
                paragraphs.append(text)
        item['type'] = 'docx'
        item['paragraph_count'] = len(paragraphs)
        item['text_preview'] = '\n'.join(paragraphs[:200])
    elif path.suffix.lower() == '.pdf':
        item['type'] = 'pdf'
        item['text_file'] = str(OUT / f'{path.stem}.txt')
    else:
        continue
    manifest.append(item)

(OUT / 'manifest.json').write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding='utf-8')
print(json.dumps(manifest, ensure_ascii=False, indent=2))
