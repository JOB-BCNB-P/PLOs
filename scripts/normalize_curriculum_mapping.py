from pathlib import Path
import csv
import re
from openpyxl import load_workbook

source = Path('/home/ubuntu/upload/Curriculummapping-2565วพบกรุงเทพพฤษภาคม65.xlsx')
out = Path('/home/ubuntu/plo-assessment-system/docs/source_analysis/curriculum_mapping_import.csv')
wb = load_workbook(source, read_only=True, data_only=True)
ws = wb.active
rows = list(ws.iter_rows(values_only=True))
header_row = rows[3]
subheader_row = rows[4]
plo_by_col = []
current_plo = ''
for value in header_row:
    text = str(value).strip() if value is not None else ''
    match = re.search(r'PLOs?\s*(\d+)', text, re.IGNORECASE)
    if match:
        current_plo = f'PLO{match.group(1)}'
    plo_by_col.append(current_plo)

records = []
current_year = ''
for row in rows[5:]:
    first = str(row[0]).strip() if row and row[0] is not None else ''
    code = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ''
    name = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ''
    credits = str(row[3]).strip() if len(row) > 3 and row[3] is not None else ''
    if 'ชั้นปีที่' in first:
        current_year = first
    if not code or not name or not any(ch.isdigit() for ch in code):
        continue
    for col in range(4, min(len(row), len(subheader_row))):
        level = str(row[col]).strip() if row[col] is not None else ''
        sub_plo = str(subheader_row[col]).strip() if subheader_row[col] is not None else ''
        if level in {'M', 'I', 'R'} and plo_by_col[col] and sub_plo:
            records.append({
                'curriculum_version': '2565',
                'year_level': current_year,
                'course_code': code,
                'course_name_th': name,
                'credits_text': credits,
                'plo_code': plo_by_col[col],
                'sub_plo_code': sub_plo,
                'mapping_level': level,
                'source_filename': source.name,
            })

with out.open('w', newline='', encoding='utf-8-sig') as f:
    fieldnames = ['curriculum_version','year_level','course_code','course_name_th','credits_text','plo_code','sub_plo_code','mapping_level','source_filename']
    writer = csv.DictWriter(f, fieldnames=fieldnames)
    writer.writeheader()
    writer.writerows(records)
print(f'rows={len(records)} courses={len({r["course_code"] for r in records})} plos={len({r["plo_code"] for r in records})} output={out}')
